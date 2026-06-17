#!/usr/bin/env python3
"""
Market Data Ingestion — Commodity Prices

Fetches commodity price data from World Bank Pink Sheet (XLS downloads)
and inserts into price_observation (PostgreSQL).

The World Bank Commodity Price Data (Pink Sheet) covers 70+ commodities
with monthly prices. This module downloads the XLS file and extracts
prices for the 8 Kokonut commodities.

Usage:
    python3 -m services.ingestion.market_data --source world_bank
    python3 -m services.ingestion.market_data --source world_bank --month 6 --year 2026
    python3 -m services.ingestion.market_data --source seed
    python3 -m services.ingestion.market_data --commodity COFFEE
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timezone, timedelta
from io import BytesIO

import requests

from ..common.logging import get_logger
from .base import get_db, log_ingestion, hash_payload, retry

logger = get_logger("ingestion.market")

# World Bank Pink Sheet XLS URL pattern
# Updated periodically; the exact URL changes each month
WORLD_BANK_BASE_URL = "https://thedocs.worldbank.org/en/1c3f894e-2c54-48b5-a79d-2e6c9c47b3af"

# Fallback URL for the latest available Pink Sheet
WORLD_BANK_FALLBACK_URL = os.environ.get(
    "WORLD_BANK_PINK_SHEET_URL",
    f"{WORLD_BANK_BASE_URL}/CMO-Historical-Data-Monthly.xlsx",
)

# Commodity seed data (fallback when live download unavailable)
COMMODITY_SEED_DATA = {
    "COFFEE": {
        "name": "Coffee (Arabica)",
        "unit": "USD/kg",
        "prices": [
            {"date": "2024-01-01", "price": 5.23},
            {"date": "2024-04-01", "price": 6.12},
            {"date": "2024-07-01", "price": 5.87},
            {"date": "2024-10-01", "price": 6.45},
            {"date": "2025-01-01", "price": 7.89},
            {"date": "2025-04-01", "price": 8.52},
        ],
    },
    "COCOA": {
        "name": "Cocoa",
        "unit": "USD/kg",
        "prices": [
            {"date": "2024-01-01", "price": 4.12},
            {"date": "2024-04-01", "price": 4.89},
            {"date": "2024-07-01", "price": 5.23},
            {"date": "2024-10-01", "price": 6.01},
            {"date": "2025-01-01", "price": 7.15},
            {"date": "2025-04-01", "price": 7.89},
        ],
    },
    "PALM_OIL": {
        "name": "Palm Oil",
        "unit": "USD/tonne",
        "prices": [
            {"date": "2024-01-01", "price": 780},
            {"date": "2024-04-01", "price": 810},
            {"date": "2024-07-01", "price": 795},
            {"date": "2024-10-01", "price": 830},
            {"date": "2025-01-01", "price": 865},
            {"date": "2025-04-01", "price": 890},
        ],
    },
    "RICE": {
        "name": "Rice",
        "unit": "USD/tonne",
        "prices": [
            {"date": "2024-01-01", "price": 520},
            {"date": "2024-04-01", "price": 535},
            {"date": "2024-07-01", "price": 515},
            {"date": "2024-10-01", "price": 540},
            {"date": "2025-01-01", "price": 555},
            {"date": "2025-04-01", "price": 570},
        ],
    },
    "MAIZE": {
        "name": "Maize",
        "unit": "USD/tonne",
        "prices": [
            {"date": "2024-01-01", "price": 195},
            {"date": "2024-04-01", "price": 205},
            {"date": "2024-07-01", "price": 198},
            {"date": "2024-10-01", "price": 210},
            {"date": "2025-01-01", "price": 218},
            {"date": "2025-04-01", "price": 225},
        ],
    },
    "SUGAR": {
        "name": "Sugar",
        "unit": "USD/kg",
        "prices": [
            {"date": "2024-01-01", "price": 0.28},
            {"date": "2024-04-01", "price": 0.31},
            {"date": "2024-07-01", "price": 0.29},
            {"date": "2024-10-01", "price": 0.33},
            {"date": "2025-01-01", "price": 0.35},
            {"date": "2025-04-01", "price": 0.37},
        ],
    },
    "TEA": {
        "name": "Tea",
        "unit": "USD/kg",
        "prices": [
            {"date": "2024-01-01", "price": 2.45},
            {"date": "2024-04-01", "price": 2.58},
            {"date": "2024-07-01", "price": 2.52},
            {"date": "2024-10-01", "price": 2.65},
            {"date": "2025-01-01", "price": 2.78},
            {"date": "2025-04-01", "price": 2.85},
        ],
    },
    "BANANA": {
        "name": "Bananas",
        "unit": "USD/kg",
        "prices": [
            {"date": "2024-01-01", "price": 1.12},
            {"date": "2024-04-01", "price": 1.18},
            {"date": "2024-07-01", "price": 1.15},
            {"date": "2024-10-01", "price": 1.22},
            {"date": "2025-01-01", "price": 1.28},
            {"date": "2025-04-01", "price": 1.32},
        ],
    },
}

CROP_ALIASES = {
    "COFFEE": ["coffee", "arabica coffee"],
    "COCOA": ["cocoa", "cacao", "trinitario cacao"],
    "PALM_OIL": ["palm", "oil palm", "palm oil"],
    "RICE": ["rice"],
    "MAIZE": ["maize", "corn"],
    "SUGAR": ["sugar", "sugarcane", "sugar cane"],
    "TEA": ["tea"],
    "BANANA": ["banana", "bananas", "plantain"],
}

# Pink Sheet column headers we look for (case-insensitive partial matches)
_PINK_SHEET_MONTHS = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december",
]


def normalize_name(value: str) -> str:
    """Normalize crop and commodity names for forgiving matching."""
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def find_crop_id(crops: list[tuple], commodity_code: str, commodity_name: str):
    """Match commodity prices to local crop names using aliases and containment."""
    candidates = [commodity_code.replace("_", " "), commodity_name]
    candidates.extend(CROP_ALIASES.get(commodity_code, []))
    normalized_candidates = [normalize_name(candidate) for candidate in candidates]

    for crop_id, crop_name in crops:
        crop_norm = normalize_name(crop_name)
        for candidate in normalized_candidates:
            if candidate and (crop_norm == candidate or candidate in crop_norm or crop_norm in candidate):
                return crop_id
    return None


def download_pink_sheet(url: str = None) -> bytes:
    """Download the World Bank Pink Sheet XLS file.

    Returns the raw bytes of the XLS file.
    Raises requests.RequestException on network errors.
    """
    target_url = url or WORLD_BANK_FALLBACK_URL
    logger.info("Downloading Pink Sheet from %s...", target_url)

    resp = requests.get(target_url, timeout=60)
    resp.raise_for_status()

    content_type = resp.headers.get("Content-Type", "")
    if "spreadsheet" not in content_type and "excel" not in content_type and "octet-stream" not in content_type:
        logger.warning("Unexpected Content-Type: %s (proceeding anyway)", content_type)

    logger.info("Downloaded %d bytes", len(resp.content))
    return resp.content


def parse_pink_sheet(xls_bytes: bytes, target_month: int = None, target_year: int = None) -> list:
    """Parse the World Bank Pink Sheet XLS and extract commodity prices.

    Returns list of dicts with:
        commodity_code, commodity_name, price_date, price_per_unit, unit
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is required for Pink Sheet parsing: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(BytesIO(xls_bytes), read_only=True, data_only=True)

    # Try common sheet names
    sheet_name = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "monthly" in name_lower or "pink" in name_lower or "commodity" in name_lower:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        logger.warning("Empty sheet: %s", sheet_name)
        return []

    # Find header row (contains "Commodity" or "Unit")
    header_idx = None
    for i, row in enumerate(rows):
        row_str = " ".join(str(c).lower() for c in row if c)
        if "commodity" in row_str or "unit" in row_str:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 0

    headers = [str(c).strip() if c else "" for c in rows[header_idx]]

    # Find month columns
    month_cols = {}
    for col_idx, header in enumerate(headers):
        header_lower = header.lower()
        for m_idx, month_name in enumerate(_PINK_SHEET_MONTHS):
            if month_name in header_lower:
                month_cols[m_idx + 1] = col_idx
                break

    # Find year column
    year_col = None
    for col_idx, header in enumerate(headers):
        header_lower = header.lower()
        if "year" in header_lower or "date" in header_lower:
            year_col = col_idx
            break

    if not month_cols:
        logger.warning("Could not identify month columns in Pink Sheet")
        return []

    # Map commodity names to our codes
    COMMODITY_NAME_MAP = {
        "coffee": "COFFEE",
        "arabica": "COFFEE",
        "cocoa": "COCOA",
        "cacao": "COCOA",
        "palm oil": "PALM_OIL",
        "rice": "RICE",
        "maize": "MAIZE",
        "corn": "MAIZE",
        "sugar": "SUGAR",
        "tea": "TEA",
        "banana": "BANANA",
        "bananas": "BANANA",
    }

    records = []
    current_year = None

    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        # Check if this row has a year
        if year_col is not None and row[year_col] is not None:
            try:
                year_val = int(row[year_col])
                if 1960 <= year_val <= 2030:
                    current_year = year_val
            except (ValueError, TypeError):
                pass

        if current_year is None:
            continue

        # First column is typically the commodity name
        commodity_name = str(row[0]).strip() if row[0] else ""
        commodity_lower = commodity_name.lower()

        commodity_code = None
        for key, code in COMMODITY_NAME_MAP.items():
            if key in commodity_lower:
                commodity_code = code
                break

        if not commodity_code:
            continue

        # Extract prices from month columns
        for month_num, col_idx in month_cols.items():
            if target_month and month_num != target_month:
                continue
            if target_year and current_year != target_year:
                continue

            if col_idx >= len(row):
                continue

            val = row[col_idx]
            if val is None:
                continue

            try:
                price = float(val)
            except (ValueError, TypeError):
                continue

            if price <= 0:
                continue

            date_str = f"{current_year}-{month_num:02d}-01"
            records.append({
                "commodity_code": commodity_code,
                "commodity_name": commodity_name,
                "price_date": date_str,
                "price_per_unit": price,
            })

    return records


def insert_price(db, record: dict) -> str:
    """Insert price observation into PostgreSQL. Returns record ID."""
    with db.cursor() as cur:
        cur.execute(
            """
            INSERT INTO price_observation
                (crop_id, commodity_code, market_name, price_date,
                 price_per_unit, unit, currency, source, source_url, metadata)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
            ON CONFLICT DO NOTHING
            RETURNING id
            """,
            (
                record.get("crop_id"),
                record["commodity_code"],
                record.get("market_name", "World Bank Pink Sheet"),
                record["price_date"],
                record["price_per_unit"],
                record["unit"],
                record.get("currency", "USD"),
                record["source"],
                record.get("source_url", ""),
                json.dumps(record.get("metadata", {})),
            ),
        )
        row = cur.fetchone()
        return str(row[0]) if row else None


def run_seed_data(commodity: str = None):
    """Seed prices from hardcoded data (fallback when live download unavailable)."""
    db = get_db()

    with db.cursor() as cur:
        cur.execute("SELECT id, name FROM crop")
        crops = cur.fetchall()

    commodities = [commodity] if commodity else list(COMMODITY_SEED_DATA.keys())

    logger.info("Seeding prices for %d commodities...", len(commodities))
    success = 0
    errors = 0

    for code in commodities:
        if code not in COMMODITY_SEED_DATA:
            logger.warning("  Unknown commodity: %s", code)
            continue

        info = COMMODITY_SEED_DATA[code]
        try:
            crop_id = find_crop_id(crops, code, info["name"])
            inserted = 0

            for price_point in info["prices"]:
                record = {
                    "crop_id": crop_id,
                    "commodity_code": code,
                    "market_name": "World Bank Pink Sheet",
                    "price_date": price_point["date"],
                    "price_per_unit": price_point["price"],
                    "unit": info["unit"],
                    "currency": "USD",
                    "source": "world_bank_pink_sheet",
                    "source_url": "https://www.worldbank.org/en/research/commodity-markets",
                    "metadata": {"source_type": "seed_data"},
                }
                insert_price(db, record)
                inserted += 1

            success += 1
            logger.info("  %s: %d price records seeded", info['name'], inserted)

        except Exception as e:
            errors += 1
            logger.error("  %s: %s", info['name'], e)

    db.commit()
    db.close()
    logger.info("Done: %d success, %d errors", success, errors)


def run_world_bank(commodity: str = None, month: int = None, year: int = None):
    """Download and parse World Bank Pink Sheet, then insert prices."""
    db = get_db()

    with db.cursor() as cur:
        cur.execute("SELECT id, name FROM crop")
        crops = cur.fetchall()

    try:
        xls_bytes = download_pink_sheet()
    except Exception as e:
        logger.error("Failed to download Pink Sheet: %s", e)
        logger.info("Falling back to seed data...")
        db.close()
        run_seed_data(commodity=commodity)
        return

    records = parse_pink_sheet(xls_bytes, target_month=month, target_year=year)

    if not records:
        logger.warning("No records extracted from Pink Sheet")
        db.close()
        return

    logger.info("Extracted %d price records from Pink Sheet", len(records))
    success = 0
    errors = 0

    for rec in records:
        if commodity and rec["commodity_code"] != commodity:
            continue

        try:
            crop_id = find_crop_id(crops, rec["commodity_code"], rec["commodity_name"])

            unit_map = {
                "COFFEE": "USD/kg",
                "COCOA": "USD/kg",
                "PALM_OIL": "USD/tonne",
                "RICE": "USD/tonne",
                "MAIZE": "USD/tonne",
                "SUGAR": "USD/kg",
                "TEA": "USD/kg",
                "BANANA": "USD/kg",
            }

            record = {
                "crop_id": crop_id,
                "commodity_code": rec["commodity_code"],
                "market_name": "World Bank Pink Sheet",
                "price_date": rec["price_date"],
                "price_per_unit": rec["price_per_unit"],
                "unit": unit_map.get(rec["commodity_code"], "USD/kg"),
                "currency": "USD",
                "source": "world_bank_pink_sheet",
                "source_url": WORLD_BANK_FALLBACK_URL,
                "metadata": {"source_type": "live_download"},
            }
            insert_price(db, record)
            success += 1

        except Exception as e:
            errors += 1
            logger.error("  %s %s: %s", rec["commodity_code"], rec["price_date"], e)

    db.commit()
    db.close()
    logger.info("Done: %d success, %d errors", success, errors)


def run(commodity: str = None, source: str = "world_bank", month: int = None, year: int = None):
    """Main ingestion entry point."""
    if source == "seed":
        run_seed_data(commodity=commodity)
    elif source == "world_bank":
        run_world_bank(commodity=commodity, month=month, year=year)
    else:
        logger.error("Unknown source: %s. Use 'world_bank' or 'seed'.", source)
        sys.exit(1)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Market data ingestion")
    parser.add_argument("--commodity", help="Specific commodity code (e.g., COFFEE)")
    parser.add_argument(
        "--source",
        default="world_bank",
        choices=["world_bank", "seed"],
        help="Data source (default: world_bank)",
    )
    parser.add_argument("--month", type=int, help="Month (1-12) to fetch from Pink Sheet")
    parser.add_argument("--year", type=int, help="Year to fetch from Pink Sheet")
    args = parser.parse_args()
    run(commodity=args.commodity, source=args.source, month=args.month, year=args.year)
